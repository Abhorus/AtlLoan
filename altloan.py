import openpyxl 
import os
from fuzzywuzzy import fuzz
import datetime #use timedelta for previous days

start_time = datetime.datetime.now()
os.chdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Training")

wb = openpyxl.load_workbook('List of commonly received Alternative Loans.xlsx')
sheet = wb.active


a = list(sheet.columns)[0]
altloanNames = []
for i in range(1,len(a)):
    altloanNames.append(a[i].value)

os.chdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Queries\Enrollment Queries\FY21")

schpwb = openpyxl.load_workbook('2021_03_03_OSF_SCHOLARSHIP_PSTD_ENROLMNT.xlsx')
#schpwb = openpyxl.load_workbook('_Altloan_Testfile_2020_11_04_OSF_SCHOLARSHIP_PSTD_ENROLMNT.xlsx')
schpsheet = schpwb.active

refcolumn = list(schpsheet.columns)[7]

os.chdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Queries\Alt Loan Query Results")
altLoanresults = openpyxl.Workbook()
resultsSheet = altLoanresults.active
resultsSheet.cell(row=1, column=1).value = 'ID'
resultsSheet.cell(row=1, column=2).value = 'Item Type'
resultsSheet.cell(row=1, column=3).value = 'Descr'
resultsSheet.cell(row=1, column=4).value = 'Item Amt'
resultsSheet.cell(row=1, column=5).value = 'Term'
resultsSheet.cell(row=1, column=6).value = 'Take Prgrs'
resultsSheet.cell(row=1, column=7).value = 'Career'
resultsSheet.cell(row=1, column=8).value = 'Ref Nbr'
resultsSheet.cell(row=1, column=9).value = 'Postd DtTm'
resultsSheet.cell(row=1, column=10).value = 'User'

####testing###
#for i, ele in enumerate(refcolumn,1):
    #print(i, ele.value) #gets reference column 
#    if schpsheet.cell(row=i, column=9).value == None or type(schpsheet.cell(row=i, column=9).value) == str: #moves on if column has nothing in it or a string
#        continue
        
#    elif schpsheet.cell(row=i, column=9).value.date() == datetime.date.today()-datetime.timedelta(1): if today's date matches previoues date on spreadsheet, gets the data from that row/column
#        print(i,schpsheet.cell(row=i, column=9).value.date())
    #else:
        #print(i, schpsheet.cell(row=i, column=9).value.date(), datetime.date.today()-datetime.timedelta(1))

################################
count = 0
if datetime.datetime.today().weekday() == 0:
    for i in altloanNames:
        for j, ele in enumerate(refcolumn):
            if schpsheet.cell(row=j+1, column=9).value == None or type(schpsheet.cell(row=j+1, column=9).value) == str: #moves on if column has nothing in it or a string
                continue
            #elif schpsheet.cell(row=j, column=9).value.date() == datetime.date.today()-datetime.timedelta(1):
            elif schpsheet.cell(row=j+1, column=9).value.date() == datetime.date.today()-datetime.timedelta(3) and fuzz.token_set_ratio(i, ele.value) > 90:
                #print(fuzz.token_set_ratio(i, ele.value),i ,ele.value)
                count += 1
                for index, element in enumerate(list(schpsheet.rows)[j]): #why not j+1 here?
                    resultsSheet.cell(row=count+1, column=index+1).value = element.value
                    print(element.value, end=" ")
                print('\n')
else:  
    for i in altloanNames:
        for j, ele in enumerate(refcolumn):
            if schpsheet.cell(row=j+1, column=9).value == None or type(schpsheet.cell(row=j+1, column=9).value) == str: #moves on if column has nothing in it or a string
                continue
            #elif schpsheet.cell(row=j, column=9).value.date() == datetime.date.today()-datetime.timedelta(1):
            #else:
            #   print(fuzz.token_set_ratio(i, ele.value),i,ele.value, schpsheet.cell(row=j+1, column=1).value) 
            elif schpsheet.cell(row=j+1, column=9).value.date() == datetime.date.today()-datetime.timedelta(1) and fuzz.token_set_ratio(i, ele.value) > 90:
                #print(fuzz.token_set_ratio(i, ele.value),i, ele.value)
                count += 1
                for index, element in enumerate(list(schpsheet.rows)[j]): #why not j+1 here?
                    resultsSheet.cell(row=count+1, column=index+1).value = element.value
                    print(element.value, end=" ")
                print('\n')

                
altLoanresults.save('Alt Loan Results_'+ str(datetime.date.today()) + '.xlsx')
print('Elapsed: ',datetime.datetime.now() - start_time )